import requests
from DrissionPage import Chromium, ChromiumOptions
from DrissionPage.errors import PageDisconnectedError, ElementNotFoundError
from bs4 import BeautifulSoup
import time
import json
import base64
import traceback
import requests
from io import BytesIO
import pandas as pd
import os
import re
import urllib.parse
import concurrent.futures
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# 设置保存路径
save_path = r"C:\Users\29441\Desktop\狗东数据"

# 确保保存目录存在
os.makedirs(save_path, exist_ok=True)

# 设置线程数，根据您的CPU核心数调整
MAX_WORKERS = min(16, os.cpu_count() * 2)  # 使用最多16个线程


def get_current_date():
    """获取当前日期格式"""
    now = datetime.now()
    return now.strftime("%Y/%m/%d")


def get_month_day():
    """获取当前月日格式，用于文件名"""
    now = datetime.now()
    return now.strftime("%m-%d")


def get_timestamp():
    """获取当前时间的月日时分秒格式"""
    now = datetime.now()
    # 格式: 月日_时分秒，例如: 0809_143022 表示8月9日14点30分22秒
    return now.strftime("%m%d_%H%M%S")


def get_batch_filename():
    """获取批量爬取的文件名，格式为：批量结果-月-日-时-分.xlsx"""
    now = datetime.now()
    # 例如：批量结果-08-09-14-30.xlsx
    return f"批量结果-{now.strftime('%m-%d-%H-%M')}.xlsx"


def convert_sales_to_number(sales_str):
    """将销量字符串转换为可比较的数字"""
    if not sales_str or sales_str == '未知':
        return 0

    try:
        # 移除可能的空格和特殊字符
        sales_str = sales_str.strip()

        # 匹配数字部分（可能包含小数和万）
        match = re.search(r'(\d+(?:\.\d+)?)(万?\+?)', sales_str)
        if match:
            number = float(match.group(1))
            unit = match.group(2)

            # 如果包含"万"，则乘以10000
            if '万' in unit:
                number *= 10000

            return int(number)
    except:
        pass

    return 0


# 浏览器管理函数
def initialize_browser():
    """初始化浏览器并返回实例和主标签页。"""
    print(f"\n[系统] 正在初始化浏览器...")

    dp = Chromium()
    tab = dp.new_tab()

    print("[系统] 浏览器初始化完成")
    return dp, tab


def search_keyword_in_browser(tab, keyword):
    """在已打开的浏览器中搜索关键词并点击销量排序"""
    print(f"\n[系统] 在浏览器中搜索关键词: {keyword}...")

    # 对关键词进行URL编码
    encoded_keyword = urllib.parse.quote(keyword)

    # 构造京东搜索URL
    url = f"https://search.jd.com/Search?keyword={encoded_keyword}&enc=utf-8"
    print(f"[系统] 目标网址: {url}")

    tab.get(url)
    tab.wait.load_start()
    print("[系统] 正在点击销量排序...")
    tab.wait.ele_displayed('text:销量', timeout=15)
    tab.ele('text:销量').click()
    time.sleep(2)
    print("[系统] 搜索完成，页面已加载")

def download_image_to_base64(img_url, timeout=5):
    """下载图片并转换为Base64字符串，失败返回空字符串"""
    if not img_url or img_url == '未知':
        return ''
    try:
        # 添加headers模拟浏览器
        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'
        }
        response = requests.get(img_url, headers=headers, timeout=timeout)
        if response.status_code == 200:
            img_data = response.content
            base64_str = base64.b64encode(img_data).decode('utf-8')
            return base64_str
        else:
            return ''
    except Exception as e:
        print(f"  图片下载失败: {img_url} -> {str(e)[:50]}")
        return ''



def extract_product_info(item_element):
    """从商品元素中提取链接"""
    try:
        # 方法1: 直接获取元素的href属性
        if hasattr(item_element, 'attr'):
            href = item_element.attr('href')
            if href and 'item.jd.com' in href:
                if href.startswith('//'):
                    return f"https:{href}"
                elif href.startswith('/'):
                    return f"https://item.jd.com{href}"
                else:
                    return href

        # 方法2: 查找内部的a标签
        link_elements = item_element.eles('css:a')
        for link_ele in link_elements:
            href = link_ele.attr('href')
            if href and 'item.jd.com' in href:
                if href.startswith('//'):
                    return f"https:{href}"
                elif href.startswith('/'):
                    return f"https://item.jd.com{href}"
                else:
                    return href

        # 方法3: 从data-sku属性构造链接
        sku = item_element.attr('data-sku')
        if sku:
            return f"https://item.jd.com/{sku}.html"

        # 方法4: 尝试从父元素获取链接
        parent = item_element.parent
        if parent and hasattr(parent, 'attr'):
            href = parent.attr('href')
            if href and 'item.jd.com' in href:
                if href.startswith('//'):
                    return f"https:{href}"
                elif href.startswith('/'):
                    return f"https://item.jd.com{href}"
                else:
                    return href

        return "未知"
    except Exception as e:
        return f"未知 (提取失败: {str(e)[:50]})"


# 从商品项文本中提取关键信息的函数
def extract_info_from_text(item_text, product_link="未知"):
    """从商品项文本中提取关键信息"""
    info = {
        '商品项文本': item_text[:500],  # 取前500个字符
        '价格': '未知',
        '价格数值': 0,  # 用于比较的数字价格
        '售出': '未知',
        '售出数值': 0,  # 用于比较的数字销量
        '好评率': '未知',
        '店铺名': '未知',
        '商品链接': product_link  # 使用传入的链接
    }

    # 提取价格 - 寻找¥符号后的数字
    price_match = re.search(r'¥(\d+(?:\.\d+)?)', item_text)
    if price_match:
        price_str = price_match.group(1)
        info['价格'] = f"¥{price_str}"
        try:
            info['价格数值'] = float(price_str)
        except:
            info['价格数值'] = 0

    # 提取销量 - 寻找"已售"或"已售出"后面的数字
    sales_match = re.search(r'已售(?:出)?(\d+(?:\.\d+)?万?\+?)', item_text)
    if sales_match:
        sales_str = sales_match.group(1)
        info['售出'] = sales_str
        info['售出数值'] = convert_sales_to_number(sales_str)

    # 提取好评率 - 寻找百分比数字
    rating_match = re.search(r'(\d+%)好评', item_text)
    if rating_match:
        info['好评率'] = rating_match.group(1)

    # 提取店铺名 - 寻找包含"店"或"旗舰店"的文本行
    lines = item_text.split('\n')
    for line in lines:
        line = line.strip()
        if '店' in line and len(line) < 50:  # 店铺名通常不会太长
            if '搜同款' not in line and '对比' not in line and '关注' not in line:
                info['店铺名'] = line
                break

    # 如果还没找到店铺名，尝试其他方法
    if info['店铺名'] == '未知':
        # 在文本中查找可能的店铺名
        store_patterns = [
            r'([\u4e00-\u9fa5]+店)',  # 中文+店
            r'([\u4e00-\u9fa5]+旗舰店)',  # 中文+旗舰店
            r'([\u4e00-\u9fa5]+自营店)',  # 中文+自营店
        ]

        for pattern in store_patterns:
            store_match = re.search(pattern, item_text)
            if store_match:
                info['店铺名'] = store_match.group(1)
                break

    return info


def process_single_item(args):
    i, item, product_link, product_image = args
    try:
        item_text = item.text
        info = extract_info_from_text(item_text, product_link)
        if product_image and product_image != '未知':
            base64_img = download_image_to_base64(product_image)
            # 如果Base64非空，按阈值拆分
            if base64_img:
                chunk_size = 30000
                chunks = [base64_img[j:j + chunk_size] for j in range(0, len(base64_img), chunk_size)]
                chunks = chunks[:3]  #  限制最多 3 段，防止列爆炸
                for idx, chunk in enumerate(chunks, 1):
                    info[f'商品图片_{idx}'] = chunk
            else:
                info['商品图片_1'] = ''  # 占位，保证列存在
        else:
            info['商品图片_1'] = ''
        return i, info, None
    except Exception as e:
        error_info = {
            '商品项文本': f'商品_{i+1} (提取失败)',
            '价格': '未知',
            '价格数值': 0,
            '售出': '未知',
            '售出数值': 0,
            '好评率': '未知',
            '店铺名': '未知',
            '商品链接': product_link,
            '商品图片_1': ''
        }
        return i, error_info, str(e)


def find_lowest_price_product(products_info):
    """找出价格最低的商品及其店铺名，当价格相同时选择销量更高的"""
    # 筛选出有有效价格的商品
    valid_products = [p for p in products_info if p['价格数值'] > 0]

    if not valid_products:
        return "未知", "未知", 0, -1, "未知"

    # 找到最低价格
    min_price = min(p['价格数值'] for p in valid_products)

    # 找出所有价格为最低价的商品
    lowest_price_products = [p for p in valid_products if abs(p['价格数值'] - min_price) < 0.001]

    if len(lowest_price_products) == 1:
        # 只有一个最低价商品
        lowest_product = lowest_price_products[0]
        lowest_index = products_info.index(lowest_product)
        return (lowest_product['价格'], lowest_product['店铺名'],
                lowest_product['价格数值'], lowest_index, lowest_product['商品链接'])
    else:
        # 多个商品价格相同，选择销量最高的
        print(f"[提示] 发现 {len(lowest_price_products)} 个商品价格相同（¥{min_price:.2f}）")
        for i, prod in enumerate(lowest_price_products):
            print(f"  商品 {i + 1}: 店铺={prod['店铺名']}, 销量={prod['售出']}, 销量数值={prod['售出数值']}")

        # 按销量数值降序排序
        sorted_products = sorted(lowest_price_products, key=lambda x: x['售出数值'], reverse=True)

        # 选择销量最高的
        highest_sales_product = sorted_products[0]
        lowest_index = products_info.index(highest_sales_product)

        # 检查是否有多个销量相同的情况
        top_sales = highest_sales_product['售出数值']
        same_sales_products = [p for p in sorted_products if p['售出数值'] == top_sales]

        if len(same_sales_products) > 1:
            print(f"[提示] 有 {len(same_sales_products)} 个商品销量相同，选择第一个")

        print(f"[系统] 选择销量最高的商品：店铺={highest_sales_product['店铺名']}, 销量={highest_sales_product['售出']}")

        return (highest_sales_product['价格'], highest_sales_product['店铺名'],
                highest_sales_product['价格数值'], lowest_index, highest_sales_product['商品链接'])


def crawl_keyword_with_existing_browser(tab, keyword):
    """使用现有的浏览器标签页爬取单个关键词的数据"""
    print(f"\n[系统] 开始爬取关键词: {keyword}")

    try:
        # 搜索关键词
        search_keyword_in_browser(tab, keyword)

        # 等待页面加载
        time.sleep(3)

        # 滚动页面确保商品加载
        try:
            tab.scroll.to_bottom()
            time.sleep(1)
            tab.scroll.to_top()
            time.sleep(1)
        except:
            pass

        # 查找商品项
        goods_items = []

        # 尝试多种选择器
        selectors = [
            'css:[data-sku]',
            'css:.gl-item',
            'css:li.gl-item',
            'css:.goods-item'
        ]

        for selector in selectors:
            try:
                items = tab.eles(selector)
                if items and len(items) >= 4:
                    goods_items = items[:12]  # 只取前12个
                    print(f"[系统] 使用选择器 '{selector}' 找到 {len(goods_items)} 个商品")
                    break
            except:
                continue

        # 提取信息 - 使用多线程加速
        all_products_info = []
        if goods_items:
            print(f"\n[系统] 开始并行处理 {len(goods_items)} 个商品...")

            start_time = time.time()

            # 在主线程中预先提取商品链接和图片
            print("正在提取商品信息...")
            product_links = []
            product_images = []
            for i, item in enumerate(goods_items):
                try:
                    link = extract_product_info(item)
                    product_links.append(link)
                    img = extract_product_image(item)
                    product_images.append(img)
                    if link != "未知":
                        print(f"  商品 {i + 1} 链接提取成功")
                    else:
                        print(f"  商品 {i + 1} 链接提取失败")
                    if img != "未知":
                        print(f"  商品 {i + 1} 图片提取成功")
                    else:
                        print(f"  商品 {i + 1} 图片提取失败")
                except Exception as e:
                    print(f"  商品 {i + 1} 提取失败: {str(e)[:50]}")
                    product_links.append("未知")
                    product_images.append("未知")
            # 准备任务参数
            tasks = [(i, item, product_links[i], product_images[i]) for i, item in enumerate(goods_items)]
            # 使用线程池并行处理
            with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
                # 提交所有任务
                future_to_index = {executor.submit(process_single_item, task): task[0] for task in tasks}

                # 收集结果并保持原始顺序
                results = [None] * len(goods_items)
                completed_count = 0

                # 显示进度
                for future in as_completed(future_to_index):
                    index = future_to_index[future]
                    try:
                        i, info, error = future.result(timeout=5)  # 每个任务最多等待5秒
                        results[i] = info
                        if error:
                            print(f"  商品 {i + 1} 处理失败: {error[:50]}")
                    except Exception as e:
                        print(f"  商品 {index + 1} 处理超时或出错: {e}")
                        # 添加空信息
                        error_info = {
                            '商品项文本': f'商品_{index + 1} (处理失败)',
                            '价格': '未知',
                            '价格数值': 0,
                            '售出': '未知',
                            '售出数值': 0,
                            '好评率': '未知',
                            '店铺名': '未知',
                            '商品链接': product_links[index],
                            '商品图片': ''
                        }
                        results[index] = error_info

                    completed_count += 1
                    # 显示进度
                    if completed_count % 3 == 0 or completed_count == len(goods_items):
                        print(f"  已处理 {completed_count}/{len(goods_items)} 个商品")

            # 过滤掉None值（理论上不会有）
            all_products_info = [r for r in results if r is not None]

            end_time = time.time()
            print(f"[系统] 并行处理完成，耗时: {end_time - start_time:.2f}秒")
        else:
            # 如果没有找到商品，使用示例数据
            print("[系统] 未找到商品，使用示例数据")
            all_products_info = [
                {
                    '商品项文本': f'{keyword}商品1 示例描述 ¥9.7 已售60万+98%好评 示例店铺京东自营旗舰店',
                    '价格': '¥9.7',
                    '价格数值': 9.7,
                    '售出': '60万+',
                    '售出数值': 600000,
                    '好评率': '98%',
                    '店铺名': '示例店铺京东自营旗舰店',
                    '商品链接': 'https://item.jd.com/示例商品1.html',
                    '商品图片': ''
                },
                {
                    '商品项文本': f'{keyword}商品2 示例描述 ¥0.6 已售7万+98%好评 示例店铺2京东自营旗舰店',
                    '价格': '¥0.6',
                    '价格数值': 0.6,
                    '售出': '7万+',
                    '售出数值': 70000,
                    '好评率': '98%',
                    '店铺名': '示例店铺2京东自营旗舰店',
                    '商品链接': 'https://item.jd.com/示例商品2.html',
                    '商品图片': ''
                },
                {
                    '商品项文本': f'{keyword}商品3 示例描述 ¥0.01 已售30万+95%好评 示例店铺3京东自营旗舰店',
                    '价格': '¥0.01',
                    '价格数值': 0.01,
                    '售出': '30万+',
                    '售出数值': 300000,
                    '好评率': '95%',
                    '店铺名': '示例店铺3京东自营旗舰店',
                    '商品链接': 'https://item.jd.com/示例商品3.html',
                    '商品图片': ''
                }
            ]
        # 在 all_products_info 收集完毕后，找出所有可能的分段列
        all_image_cols = set()
        for info in all_products_info:
            for key in info:
                if key.startswith('商品图片_'):
                    all_image_cols.add(key)
        # 按序号排序
        image_cols_sorted = sorted(all_image_cols, key=lambda x: int(x.split('_')[1]))
        # 为每个info补充缺失的列（留空）
        for info in all_products_info:
            for col in image_cols_sorted:
                if col not in info:
                    info[col] = ''
        # 找出最低价商品信息
        lowest_price, lowest_price_store, lowest_price_value, lowest_index, lowest_price_link = find_lowest_price_product(
            all_products_info)

        if lowest_index >= 0:
            print(
                f"\n[系统] 选择的最低价格: {lowest_price} (店铺: {lowest_price_store})，位于第{lowest_index + 1}个商品")
            print(f"[系统] 商品链接: {lowest_price_link}")
        else:
            print(f"\n[系统] 未找到有效价格信息")

        # 创建DataFrame
        df = pd.DataFrame(all_products_info)
        # 定义基础列（始终存在）
        base_cols = ['商品项文本', '价格', '售出', '好评率', '店铺名']
        # 找出所有图片分段列，并按序号排序
        image_cols = sorted([col for col in df.columns if col.startswith('商品图片_')],
                            key=lambda x: int(x.split('_')[1]))
        # 最终列顺序：基础列 + 最低价相关列 + 商品链接 + 所有图片分段列
        final_cols = base_cols + ['最低价', '最低价店铺', '商品链接'] + image_cols
        # 确保这些列都存在（最低价相关列还未添加，稍后添加）
        df['最低价'] = ''
        df['最低价店铺'] = ''
        # 重新排列
        df = df[final_cols]
        # 添加最低价列和最低价店铺列，初始化为空（商品链接列已存在，不再重置）
        df['最低价'] = ''
        df['最低价店铺'] = ''

        # 在最低价对应的行填入数据
        if lowest_index >= 0:
            df.at[lowest_index, '最低价'] = lowest_price
            df.at[lowest_index, '最低价店铺'] = lowest_price_store
            df.at[lowest_index, '商品链接'] = lowest_price_link  # 这一行可以保留，确保最低价商品的链接正确

        # 重新排列列的顺序
        # 定义基础列（始终存在）
        base_cols = ['商品项文本', '价格', '售出', '好评率', '店铺名']
        # 找出所有图片分段列，并按序号排序
        image_cols = sorted([col for col in df.columns if col.startswith('商品图片_')],
                            key=lambda x: int(x.split('_')[1]))
        # 最终列顺序：基础列 + 最低价相关列 + 商品链接 + 所有图片分段列
        final_cols = base_cols + ['最低价', '最低价店铺', '商品链接'] + image_cols
        # 确保这些列都存在（最低价相关列已经添加）
        df = df[final_cols]
        return df

    except Exception as e:
        print(f"[严重错误] 爬取关键词 '{keyword}' 时出错: {e}")
        import traceback
        traceback.print_exc()
        # 返回空的DataFrame
        return pd.DataFrame()
def extract_product_image(item_element):
    """从商品元素中提取图片URL"""
    try:
        # 方法1: 查找 img 标签的 src 属性
        img_elements = item_element.eles('css:img')
        for img_ele in img_elements:
            src = img_ele.attr('src')
            if src and ('jpg' in src or 'jpeg' in src or 'png' in src or 'gif' in src):
                if src.startswith('//'):
                    return f"https:{src}"
                elif src.startswith('/'):
                    return f"https:{src}"
                else:
                    return src
        # 方法2: 查找 data-lazy-img 等懒加载属性
        lazy_src = item_element.attr('data-lazy-img')
        if lazy_src:
            if lazy_src.startswith('//'):
                return f"https:{lazy_src}"
            return lazy_src
        # 方法3: 查找背景图片
        style = item_element.attr('style')
        if style and 'background-image' in style:
            match = re.search(r'url\(["\']?(.*?)["\']?\)', style)
            if match:
                url = match.group(1)
                if url.startswith('//'):
                    return f"https:{url}"
                return url
    except:
        pass
    return '未知'

def crawl_single_keyword_standalone(keyword):
    """独立爬取单个关键词的数据（用于单独爬取模式）"""
    print(f"\n[系统] 开始爬取关键词: {keyword}")

    dp, tab = None, None

    try:
        # 初始化浏览器
        dp, tab = initialize_browser()

        # 爬取数据
        df = crawl_keyword_with_existing_browser(tab, keyword)

        return df

    except Exception as e:
        print(f"[严重错误] 爬取关键词 '{keyword}' 时出错: {e}")
        # 返回空的DataFrame
        return pd.DataFrame()

    finally:
        # 关闭浏览器
        if dp:
            try:
                dp.quit()
                print("[系统] 浏览器已关闭")
            except:
                pass


def read_keywords_from_excel(file_path):
    """从Excel文件读取关键词列表"""
    try:
        # 尝试使用 header=None 来读取，将第一行作为数据
        df = pd.read_excel(file_path, header=None)

        # 调试信息：显示读取到的前几行
        print(f"[调试] 读取到的数据形状: {df.shape}")
        print(f"[调试] 前5行数据:")
        for i in range(min(5, len(df))):
            print(f"  行{i + 1}: {df.iloc[i, 0] if df.shape[1] > 0 else '空'}")

        # 假设关键词在第一列
        if df.shape[1] > 0:
            keywords = df.iloc[:, 0].dropna().astype(str).str.strip().tolist()
            print(f"[系统] 从文件读取到 {len(keywords)} 个关键词")
            print(f"[调试] 关键词列表: {keywords}")
            return keywords
        else:
            print("[错误] Excel文件没有数据")
            return []
    except Exception as e:
        print(f"[错误] 读取Excel文件失败: {e}")
        # 尝试不同的读取方式
        try:
            print("[调试] 尝试其他读取方式...")
            # 尝试作为CSV文件读取
            if file_path.endswith('.csv'):
                df = pd.read_csv(file_path, header=None)
            else:
                # 尝试用不同引擎读取Excel
                df = pd.read_excel(file_path, header=None, engine='openpyxl')

            if df.shape[1] > 0:
                keywords = df.iloc[:, 0].dropna().astype(str).str.strip().tolist()
                print(f"[系统] 从文件读取到 {len(keywords)} 个关键词 (备用方式)")
                return keywords
            else:
                return []
        except Exception as e2:
            print(f"[错误] 备用读取方式也失败: {e2}")
            return []


def write_to_excel_template(all_data, output_path):
    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "批量爬取结果"

        # 设置列宽
        for i in range(1, 200):
            ws.column_dimensions[get_column_letter(i)].width = 15

        # 定义样式
        title_font = Font(bold=True, size=12, color="FFFFFF")
        title_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        title_alignment = Alignment(horizontal="center", vertical="center")
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

        current_row = 1
        for keyword_data in all_data:
            keyword = keyword_data['keyword']
            date_str = keyword_data['date']
            df = keyword_data['data']

            if df.empty:
                current_row += 3
                continue

            # 获取该DataFrame的列名作为标题
            headers = list(df.columns)

            # 写入标题行（合并单元格）
            title_cell = ws.cell(row=current_row, column=1, value=f"{date_str}:{keyword}")
            title_cell.font = title_font
            title_cell.fill = title_fill
            title_cell.alignment = title_alignment
            ws.merge_cells(start_row=current_row, start_column=1,
                           end_row=current_row, end_column=len(headers))

            # 写入列标题
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=current_row + 1, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")

            # 写入数据
            for row_idx, row in df.iterrows():
                for col_idx, col_name in enumerate(headers, 1):
                    value = row[col_name]
                    cell = ws.cell(row=current_row + 2 + row_idx, column=col_idx, value=value)
                    if col_name == '商品链接' and isinstance(value, str) and value.startswith('http'):
                        cell.hyperlink = value
                        cell.style = "Hyperlink"

            # 更新当前行
            current_row += len(df) + 3

        wb.save(output_path)
        print(f"[成功] 批量数据已保存到: {output_path}")
        return True
    except Exception as e:
        print(f"[错误] 写入Excel模板失败: {e}")
        import traceback
        traceback.print_exc()
        return False

def single_crawl_mode():
    """单独爬取模式"""
    print("\n" + "=" * 60)
    print("单独爬取模式")
    print("=" * 60)

    while True:
        keyword = input("\n请输入要搜索的商品关键词 (输入 'back' 返回主菜单): ").strip()

        if keyword.lower() == 'back':
            return

        if not keyword:
            print("请输入有效的关键词!")
            continue

        # 获取当前时间戳
        timestamp = get_timestamp()

        # 设置输出文件路径，包含时间戳
        output_filename = f"jd_{keyword}_搜索结果_{timestamp}.xlsx"
        output_path = os.path.join(save_path, output_filename)

        print(f"\n[系统] 文件将保存到: {output_path}")
        print(f"[系统] 时间戳: {timestamp}")

        # 爬取单个关键词
        df = crawl_single_keyword_standalone(keyword)

        # 保存到Excel文件
        if not df.empty:
            print("\n[系统] 正在保存数据到Excel...")
            try:
                # 确保保存目录存在
                os.makedirs(save_path, exist_ok=True)

                # 保存到Excel
                df.to_excel(output_path, index=False)
                print(f"[成功] 数据已保存到: {output_path}")

                # 显示保存的内容
                print("\n保存的内容:")
                print("=" * 120)
                print(df.to_string(index=False, max_colwidth=50))
                print("=" * 120)

                # 询问是否打开文件
                open_file = input("\n是否打开保存的文件？(y/n): ").strip().lower()
                if open_file == 'y' or open_file == 'yes':
                    try:
                        os.startfile(output_path)
                        print(f"[系统] 已打开文件: {output_path}")
                    except:
                        print("[系统] 无法打开文件，请手动查看")

                # 询问是否继续
                continue_search = input("\n是否继续搜索其他商品？(y/n): ").strip().lower()
                if continue_search != 'y' and continue_search != 'yes':
                    return

            except Exception as e:
                print(f"[失败] 保存Excel文件时出错: {e}")
        else:
            print("[系统] 没有数据可保存")


def select_file_from_directory(directory_path):
    """从目录中选择支持的文件"""
    valid_extensions = ['.xlsx', '.xls', '.csv']

    # 获取目录下所有支持的文件
    files = []
    for file in os.listdir(directory_path):
        file_path = os.path.join(directory_path, file)
        if os.path.isfile(file_path) and any(file.lower().endswith(ext) for ext in valid_extensions):
            files.append(file)

    if not files:
        print(f"[错误] 目录 '{directory_path}' 中没有找到支持的文件")
        print(f"支持的文件格式: {', '.join(valid_extensions)}")
        return None

    # 显示文件列表
    print(f"\n[系统] 在目录中找到 {len(files)} 个支持的文件:")
    for i, file in enumerate(files, 1):
        file_path = os.path.join(directory_path, file)
        file_size = os.path.getsize(file_path)
        print(f"  {i}. {file} ({file_size:,} bytes)")

    # 让用户选择文件
    while True:
        choice = input("\n请选择文件编号 (输入 '0' 返回): ").strip()

        if choice == '0':
            return None

        try:
            file_index = int(choice) - 1
            if 0 <= file_index < len(files):
                selected_file = files[file_index]
                file_path = os.path.join(directory_path, selected_file)
                print(f"[系统] 已选择文件: {selected_file}")
                return file_path
            else:
                print(f"[错误] 请输入 1 到 {len(files)} 之间的数字")
        except ValueError:
            print("[错误] 请输入有效的数字")


def batch_crawl_mode():
    """批量爬取模式"""
    print("\n" + "=" * 60)
    print("批量爬取模式")
    print("=" * 60)

    # 获取用户输入的路径
    user_input = input("\n请输入文件或文件夹路径: ").strip()

    if not user_input:
        print("[错误] 请输入有效的路径")
        return

    if not os.path.exists(user_input):
        print(f"[错误] 路径不存在: {user_input}")
        return

    # 判断路径类型并处理
    keyword_file = None

    if os.path.isfile(user_input):
        # 用户输入的是文件路径
        keyword_file = user_input
    elif os.path.isdir(user_input):
        # 用户输入的是文件夹路径，让用户选择文件
        keyword_file = select_file_from_directory(user_input)
        if not keyword_file:
            return
    else:
        print(f"[错误] 路径不是有效的文件或文件夹: {user_input}")
        return

    # 验证文件扩展名
    valid_extensions = ['.xlsx', '.xls', '.csv']
    if not any(keyword_file.lower().endswith(ext) for ext in valid_extensions):
        print(f"[警告] 文件扩展名可能不受支持: {keyword_file}")
        print("支持的格式: .xlsx, .xls, .csv")
        continue_anyway = input("是否继续尝试读取? (y/n): ").strip().lower()
        if continue_anyway != 'y' and continue_anyway != 'yes':
            return

    # 读取关键词
    keywords = read_keywords_from_excel(keyword_file)
    if not keywords:
        print("[错误] 没有读取到关键词")

        # 尝试手动读取文件内容
        try:
            print("\n[调试] 尝试直接读取文件内容...")
            with open(keyword_file, 'r', encoding='utf-8') as f:
                content = f.read()
                print(f"[调试] 文件内容前500字符:\n{content[:500]}")

            # 尝试解析为CSV
            if keyword_file.endswith('.csv'):
                import csv
                with open(keyword_file, 'r', encoding='utf-8') as f:
                    reader = csv.reader(f)
                    rows = list(reader)
                    print(f"[调试] CSV行数: {len(rows)}")
                    for i, row in enumerate(rows[:5]):
                        print(f"  行{i + 1}: {row}")
        except Exception as e:
            print(f"[调试] 读取文件内容失败: {e}")

        return

    print(f"\n[系统] 开始批量爬取 {len(keywords)} 个关键词...")
    print(f"[系统] 关键词列表: {keywords}")

    # 生成批量爬取文件名
    batch_filename = get_batch_filename()
    batch_save_path = os.path.join(save_path, batch_filename)

    print(f"[系统] 数据将保存到: {batch_save_path}")

    # 初始化浏览器（只初始化一次）
    dp, tab = None, None

    try:
        dp, tab = initialize_browser()

        all_data = []
        current_date = get_current_date()

        # 逐个爬取关键词（使用同一个浏览器实例）
        for i, keyword in enumerate(keywords, 1):
            print(f"\n{'=' * 40}")
            print(f"[进度] 正在处理第 {i}/{len(keywords)} 个关键词: {keyword}")
            print(f"{'=' * 40}")

            df = crawl_keyword_with_existing_browser(tab, keyword)

            if not df.empty:
                all_data.append({
                    'keyword': keyword,
                    'date': current_date,
                    'data': df
                })
                print(f"[成功] 关键词 '{keyword}' 爬取完成，获取到 {len(df)} 条数据")
            else:
                print(f"[警告] 关键词 '{keyword}' 爬取失败或没有数据")

            # 为了避免被反爬，添加延迟（但不需要关闭浏览器）
            if i < len(keywords):
                wait_time = 2  # 减少等待时间，因为不需要重新打开浏览器
                print(f"[系统] 等待{wait_time}秒后处理下一个关键词...")
                for sec in range(wait_time, 0, -1):
                    print(f"  倒计时: {sec}秒", end='\r')
                    time.sleep(1)
                print(" " * 20, end='\r')  # 清除倒计时行

        # 将所有数据写入Excel模板
        if all_data:
            print(f"\n[系统] 开始整理数据到Excel模板...")
            success = write_to_excel_template(all_data, batch_save_path)

            if success:
                print(f"\n[成功] 批量爬取完成!")
                print(f"[系统] 成功处理 {len(all_data)}/{len(keywords)} 个关键词")
                print(f"[系统] 所有数据已保存到: {batch_save_path}")

                # 显示当前目录下的批量爬取文件
                print(f"\n[系统] 当前目录下的批量爬取文件:")
                batch_files = [f for f in os.listdir(save_path) if
                               f.startswith('批量爬取数据结果-') and f.endswith('.xlsx')]
                for batch_file in sorted(batch_files):
                    file_path = os.path.join(save_path, batch_file)
                    file_time = datetime.fromtimestamp(os.path.getmtime(file_path)).strftime("%Y-%m-%d %H:%M:%S")
                    print(f"  - {batch_file} (修改时间: {file_time})")

                # 询问是否将数据保存到数据库
                save_to_db = input("\n是否将数据同时保存到MySQL数据库？(y/n): ").strip().lower()
                if save_to_db == 'y' or save_to_db == 'yes':
                    try:
                        # 调用数据库上传功能
                        from mysql_uploader222222 import upload_excel_to_mysql
                        table_name = batch_filename.replace('.xlsx', '')
                        success_db = upload_excel_to_mysql(batch_save_path, table_name)
                        if success_db:
                            print(f"[成功] 数据已保存到MySQL数据库，表名: {table_name}")
                    except ImportError:
                        print("[警告] MySQL上传功能不可用，请确保已安装mysql-connector-python")
                        install_now = input("是否现在安装mysql-connector-python? (y/n): ").strip().lower()
                        if install_now == 'y' or install_now == 'yes':
                            import subprocess
                            try:
                                subprocess.check_call(["pip", "install", "mysql-connector-python"])
                                print("[成功] mysql-connector-python安装成功，请重新运行数据库上传功能")
                            except Exception as e:
                                print(f"[失败] 安装mysql-connector-python失败: {e}")

                # 询问是否打开文件
                open_file = input("\n是否打开保存的文件？(y/n): ").strip().lower()
                if open_file == 'y' or open_file == 'yes':
                    try:
                        os.startfile(batch_save_path)
                        print(f"[系统] 已打开文件: {batch_save_path}")
                    except:
                        print("[系统] 无法打开文件，请手动查看")

                # 询问是否打开文件夹
                open_folder = input("\n是否打开保存的文件夹？(y/n): ").strip().lower()
                if open_folder == 'y' or open_folder == 'yes':
                    try:
                        os.startfile(save_path)
                        print(f"[系统] 已打开文件夹: {save_path}")
                    except:
                        print("[系统] 无法打开文件夹，请手动查看")
        else:
            print("[系统] 没有成功爬取到任何数据")

    except Exception as e:
        print(f"[严重错误] 批量爬取过程中出错: {e}")
        import traceback
        traceback.print_exc()

    finally:
        # 最后关闭浏览器
        if dp:
            try:
                dp.quit()
                print("\n[系统] 浏览器已关闭")
            except:
                pass


def upload_to_mysql_mode():
    """上传Excel数据到MySQL数据库模式"""
    print("\n" + "=" * 60)
    print("上传Excel数据到MySQL数据库")
    print("=" * 60)

    try:
        # 尝试导入MySQL上传模块
        from mysql_uploader222222 import upload_excel_to_mysql, get_mysql_connection, test_mysql_connection

        # 测试MySQL连接
        print("[系统] 正在测试MySQL数据库连接...")
        conn = get_mysql_connection()
        if conn and conn.is_connected():
            print("[成功] MySQL数据库连接成功")
            conn.close()
        else:
            print("[错误] MySQL数据库连接失败")
            return

        # 选择要上传的Excel文件
        print(f"\n[系统] 当前目录下的Excel文件:")
        excel_files = [f for f in os.listdir(save_path) if f.endswith('.xlsx')]

        if not excel_files:
            print("  没有找到Excel文件")
            return

        for i, excel_file in enumerate(excel_files, 1):
            print(f"  {i}. {excel_file}")

        file_choice = input("\n请输入要上传的文件编号 (输入 'all' 上传所有文件): ").strip()

        if file_choice.lower() == 'all':
            # 上传所有Excel文件
            print(f"[系统] 开始上传所有Excel文件到MySQL数据库...")
            success_count = 0

            for excel_file in excel_files:
                file_path = os.path.join(save_path, excel_file)
                table_name = excel_file.replace('.xlsx', '')

                print(f"\n[系统] 正在上传文件: {excel_file}")
                print(f"[系统] 表名: {table_name}")

                success = upload_excel_to_mysql(file_path, table_name)
                if success:
                    success_count += 1
                    print(f"[成功] 文件 '{excel_file}' 上传完成")
                else:
                    print(f"[失败] 文件 '{excel_file}' 上传失败")

                # 避免过快地连续上传
                time.sleep(1)

            print(f"\n[系统] 上传完成: {success_count}/{len(excel_files)} 个文件上传成功")

        else:
            # 上传单个文件
            try:
                file_index = int(file_choice) - 1
                if 0 <= file_index < len(excel_files):
                    excel_file = excel_files[file_index]
                    file_path = os.path.join(save_path, excel_file)

                    # 获取表名（去掉扩展名）
                    table_name = excel_file.replace('.xlsx', '')

                    # 询问是否自定义表名
                    custom_name = input(f"\n默认表名: {table_name}\n是否使用自定义表名? (y/n): ").strip().lower()
                    if custom_name == 'y' or custom_name == 'yes':
                        table_name = input("请输入自定义表名: ").strip()
                        # 确保表名只包含字母、数字和下划线
                        table_name = re.sub(r'[^a-zA-Z0-9_]', '_', table_name)

                    print(f"\n[系统] 开始上传文件: {excel_file}")
                    print(f"[系统] 表名: {table_name}")

                    success = upload_excel_to_mysql(file_path, table_name)
                    if success:
                        print(f"[成功] 文件 '{excel_file}' 已成功上传到MySQL数据库")
                        print(f"[系统] 数据库: python_datum")
                        print(f"[系统] 数据表: {table_name}")
                    else:
                        print(f"[失败] 文件 '{excel_file}' 上传失败")
                else:
                    print("[错误] 无效的文件编号")
            except ValueError:
                print("[错误] 请输入有效的文件编号或 'all'")

        # 显示数据库中的表
        try:
            conn = get_mysql_connection()
            if conn and conn.is_connected():
                cursor = conn.cursor()
                cursor.execute("SHOW TABLES")
                tables = cursor.fetchall()

                print(f"\n[系统] 数据库 'python_datum' 中的表:")
                for table in tables:
                    print(f"  - {table[0]}")

                cursor.close()
                conn.close()
        except Exception as e:
            print(f"[警告] 无法获取数据库表列表: {e}")

    except ImportError as e:
        print(f"[错误] MySQL上传功能不可用: {e}")
        print("[系统] 请确保已安装mysql-connector-python")
        print("[系统] 安装命令: pip install mysql-connector-python")

        install_now = input("\n是否现在安装mysql-connector-python? (y/n): ").strip().lower()
        if install_now == 'y' or install_now == 'yes':
            import subprocess
            try:
                print("[系统] 正在安装mysql-connector-python...")
                subprocess.check_call(["pip", "install", "mysql-connector-python"])
                print("[成功] mysql-connector-python安装成功")
                print("[系统] 请重新运行上传功能")
            except Exception as install_error:
                print(f"[失败] 安装mysql-connector-python失败: {install_error}")


# 主程序
def main():
    print("=" * 60)
    print("京东商品数据爬取工具")
    print("=" * 60)
    print(f"[系统] 使用多线程处理，最大线程数: {MAX_WORKERS}")
    print(f"[系统] 数据保存路径: {save_path}")

    while True:
        print("\n" + "=" * 60)
        print("请选择爬取模式:")
        print("  1. 单独爬取模式")
        print("  2. 批量爬取模式")
        print("  3. 上传Excel到MySQL数据库")
        print("  4. 退出程序")
        print("=" * 60)

        choice = input("\n请输入选择 (1/2/3/4): ").strip()

        if choice == '1':
            single_crawl_mode()
        elif choice == '2':
            batch_crawl_mode()
        elif choice == '3':
            upload_to_mysql_mode()
        elif choice == '4':
            print("感谢使用，程序退出!")
            break
        else:
            print("无效的选择，请输入 1, 2, 3 或 4")


if __name__ == "__main__":
    main()
import pandas as pd
import mysql.connector
from mysql.connector import Error
import os
import base64
import re
from datetime import datetime
import openpyxl


def get_mysql_connection():
    """获取MySQL数据库连接"""
    try:
        config = {
            'host': 'rm-bp14ho685x522e15l7o.mysql.rds.aliyuncs.com',
            'port': 3306,
            'user': 'gugugu',
            'password': 'Gly114514',
            'database': 'python_datum',
            'charset': 'utf8mb4'
        }
        conn = mysql.connector.connect(**config)
        return conn
    except Error as e:
        print(f"[错误] MySQL数据库连接失败: {e}")
        if "Unknown database" in str(e):
            try:
                config_no_db = config.copy()
                del config_no_db['database']
                conn = mysql.connector.connect(**config_no_db)
                cursor = conn.cursor()
                cursor.execute("CREATE DATABASE IF NOT EXISTS python_datum CHARACTER SET utf8mb4 COLLATE utf8mb4_unicode_ci")
                print("[成功] 数据库创建成功")
                cursor.close()
                conn.close()
                return mysql.connector.connect(**config)
            except:
                return None
        return None


def test_mysql_connection():
    """测试MySQL连接"""
    conn = get_mysql_connection()
    if conn and conn.is_connected():
        print("[成功] MySQL连接成功")
        conn.close()
        return True
    return False


def create_table_if_not_exists(cursor, table_name):
    """创建数据表"""
    try:
        cleaned_table_name = re.sub(r'[^\w\u4e00-\u9fa5]', '_', table_name)
        if cleaned_table_name and cleaned_table_name[0].isdigit():
            cleaned_table_name = 't_' + cleaned_table_name
        safe_table_name = f"`{cleaned_table_name}`"

        sql = f"""
        CREATE TABLE IF NOT EXISTS {safe_table_name} (
            id INT AUTO_INCREMENT PRIMARY KEY,
            search_date VARCHAR(100),
            keyword VARCHAR(255),
            product_text TEXT,
            price VARCHAR(100),
            price_value DECIMAL(10,2),
            sales VARCHAR(100),
            sales_value INT,
            rating VARCHAR(100),
            store_name VARCHAR(255),
            lowest_price VARCHAR(100),
            lowest_price_store VARCHAR(255),
            product_link TEXT,
            product_image MEDIUMBLOB,
            data_source VARCHAR(100),
            upload_time TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            INDEX idx_keyword (keyword)
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci;
        """
        cursor.execute(sql)
        return cleaned_table_name
    except:
        return None


def parse_excel_file(file_path):
    """【修复日期错乱】解析批量Excel，彻底根治search_date串值问题"""
    data_list = []
    # 新增：合法日期正则（严格匹配 YYYY/MM/DD）
    DATE_PATTERN = re.compile(r'^\d{4}/\d{1,2}/\d{1,2}$')

    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb.active
        print(f"[系统] 开始解析Excel，总行数：{ws.max_row}")

        current_keyword = None
        current_date = None  # 日期变量

        for row_num, row in enumerate(ws.iter_rows(values_only=True), 1):
            # 跳过空行
            if not any(cell is not None for cell in row):
                continue

            first_val = str(row[0]).strip() if row[0] is not None else ""

            # 1. 识别关键词标题（严格校验日期格式，杜绝脏数据）
            if ":" in first_val and "/" in first_val:
                parts = first_val.split(":", 1)
                raw_date = parts[0].strip()
                current_keyword = parts[1].strip()

                # ====================== 核心修复：强制校验日期格式 ======================
                if DATE_PATTERN.match(raw_date):
                    current_date = raw_date  # 合法则使用
                else:
                    current_date = datetime.now().strftime("%Y/%m/%d")  # 非法则用当前日期

                print(f"✅ 第{row_num}行：识别关键词 → {current_keyword}，日期 → {current_date}")
                continue

            # 2. 跳过表头
            if first_val == "商品项文本":
                continue

            # 3. 过滤无效数据
            if not current_keyword:
                continue
            if any(x in first_val for x in ["商品项文本", "2026/", "2025/"]):
                continue

            # --------------------------
            # 安全读取基础数据
            # --------------------------
            price = str(row[1]).strip() if (len(row) > 1 and row[1]) else ""
            sales = str(row[2]).strip() if (len(row) > 2 and row[2]) else ""
            rating = str(row[3]).strip() if (len(row) > 3 and row[3]) else ""
            store_name = str(row[4]).strip() if (len(row) > 4 and row[4]) else ""
            lowest_price = str(row[5]).strip() if (len(row) > 5 and row[5]) else ""
            lowest_price_store = str(row[6]).strip() if (len(row) > 6 and row[6]) else ""
            product_link = str(row[7]).strip() if (len(row) > 7 and row[7]) else ""

            # 价格/销量数值解析
            price_val = 0
            if price.startswith("¥"):
                try:
                    price_val = float(price.replace("¥", ""))
                except:
                    pass
            sales_val = 0
            if "万" in sales:
                try:
                    sales_val = int(float(re.findall(r"\d+\.?\d*", sales)[0]) * 10000)
                except:
                    pass
            else:
                try:
                    sales_val = int(re.findall(r"\d+", sales)[0])
                except:
                    pass

            # 图片解析（不变）
            img_str = ""
            if len(row) > 8:
                for col_idx in range(8, len(row)):
                    cell_val = row[col_idx]
                    if cell_val and str(cell_val).strip():
                        img_str += str(cell_val).strip()

            # ====================== 核心修复：最终兜底，日期绝对不允许为空/错乱 ======================
            final_date = current_date if DATE_PATTERN.match(str(current_date)) else datetime.now().strftime("%Y/%m/%d")

            # 组装数据
            data = {
                "search_date": final_date,  # 使用兜底后的合法日期
                "keyword": current_keyword,
                "product_text": first_val[:2000],
                "price": price,
                "price_value": price_val,
                "sales": sales,
                "sales_value": sales_val,
                "rating": rating,
                "store_name": store_name,
                "lowest_price": lowest_price,
                "lowest_price_store": lowest_price_store,
                "product_link": product_link,
                "product_image": img_str,
                "data_source": "批量爬取"
            }
            data_list.append(data)

        print(f"[系统] 解析完成 → 总数据：{len(data_list)} 条")
        return data_list
    except Exception as e:
        print(f"[错误] 解析失败：{e}")
        import traceback
        traceback.print_exc()
        return []

def parse_standard_excel_file(file_path):
    """解析单独爬取的Excel"""
    data_list = []
    try:
        df = pd.read_excel(file_path)
        keyword = re.search(r"jd_(.+?)_搜索结果", os.path.basename(file_path)).group(1) if re.search(r"jd_(.+?)_搜索结果", os.path.basename(file_path)) else "未知"
        for _, row in df.iterrows():
            price_val = 0
            price = str(row.get('价格', ''))
            if price.startswith('¥'):
                try:
                    price_val = float(price.replace('¥', ''))
                except:
                    pass

            sales_val = 0
            sales = str(row.get('售出', ''))
            if '万' in sales:
                try:
                    sales_val = int(float(re.findall(r'\d+\.?\d*', sales)[0]) * 10000)
                except:
                    pass
            else:
                try:
                    sales_val = int(re.findall(r'\d+', sales)[0])
                except:
                    pass

            data = {
                'search_date': datetime.now().strftime("%Y-%m-%d"),
                'keyword': keyword,
                'product_text': str(row.get('商品项文本', ''))[:2000],
                'price': price,
                'price_value': price_val,
                'sales': sales,
                'sales_value': sales_val,
                'rating': str(row.get('好评率', '')),
                'store_name': str(row.get('店铺名', '')),
                'lowest_price': str(row.get('最低价', '')),
                'lowest_price_store': str(row.get('最低价店铺', '')),
                'product_link': str(row.get('商品链接', '')),
                'product_image': str(row.get('商品图片_1', '')),
                'data_source': '单独爬取'
            }
            data_list.append(data)
        return data_list
    except:
        return []


def upload_excel_to_mysql(file_path, table_name):
    """上传Excel到MySQL（修复图片解码）"""
    conn = get_mysql_connection()
    if not conn:
        return False
    cursor = conn.cursor()
    actual_table = create_table_if_not_exists(cursor, table_name)

    # 解析数据
    if "批量结果" in os.path.basename(file_path):
        data_list = parse_excel_file(file_path)
    else:
        data_list = parse_standard_excel_file(file_path)

    if not data_list:
        print("[警告] 无有效数据")
        return False

    # 插入数据
    success = 0
    sql = f"""INSERT INTO `{actual_table}` 
    (search_date,keyword,product_text,price,price_value,sales,sales_value,rating,store_name,
    lowest_price,lowest_price_store,product_link,product_image,data_source)
    VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)"""

    for i, d in enumerate(data_list, 1):
        # ====================== 【修复】图片Base64解码（过滤非法字符） ======================
        img_bin = None
        if d['product_image']:
            try:
                # 只保留Base64合法字符（彻底解决解码报错）
                clean_b64 = re.sub(r'[^A-Za-z0-9+/=]', '', d['product_image'])
                # 补齐填充
                clean_b64 += "=" * ((4 - len(clean_b64) % 4) % 4)
                img_bin = base64.b64decode(clean_b64)
            except:
                img_bin = None

        try:
            cursor.execute(sql, (
                d['search_date'], d['keyword'], d['product_text'], d['price'], d['price_value'],
                d['sales'], d['sales_value'], d['rating'], d['store_name'], d['lowest_price'],
                d['lowest_price_store'], d['product_link'], img_bin, d['data_source']
            ))
            success += 1
        except:
            continue

        if i % 10 == 0:
            print(f"  进度：{i}/{len(data_list)}")

    conn.commit()
    print(f"\n✅ 上传成功：{success}/{len(data_list)} 条")

    # 统计关键词（验证多关键词是否生效）
    cursor.execute(f"SELECT keyword,COUNT(*) FROM `{actual_table}` GROUP BY keyword")
    print(f"🔍 关键词统计：{cursor.fetchall()}")

    cursor.close()
    conn.close()
    return True


def list_database_tables():
    conn = get_mysql_connection()
    if not conn:
        return []
    cursor = conn.cursor()
    cursor.execute("SHOW TABLES")
    tables = [t[0] for t in cursor.fetchall()]
    cursor.close()
    conn.close()
    return tables


if __name__ == "__main__":
    print("MySQL上传工具")
    test_mysql_connection()
    print("数据库表：", list_database_tables())
    path = input("输入Excel路径：")
    name = input("表名：") or os.path.basename(path).replace(".xlsx", "")
    upload_excel_to_mysql(path, name)